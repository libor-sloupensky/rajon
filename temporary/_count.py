"""Spočítá skutečné (neprázdné) datové řádky per sheet a extrahuje všechny unikátní sloupce."""
import openpyxl
from pathlib import Path
import io

HERE = Path(__file__).parent
files = sorted(HERE.glob("*.xlsx"))

out = io.StringIO()
totals = {}

for path in files:
    out.write(f"\n### {path.name}\n")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    file_total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        nonempty = 0
        sample_cols = set()
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # hledat řádek s hlavičkou (obsahuje "nazev" / "Akce" / "Datum" / "datum")
            if header is None:
                lower = [str(v).strip().lower() if v else "" for v in row]
                if any(h in lower for h in ("nazev", "akce", "název", "akce ")):
                    header = [str(v).strip() if v else "" for v in row]
                    continue
            # datový řádek — je aspoň 1 buňka non-empty a první sloupec něco obsahuje
            vals = [v for v in row if v not in (None, "", " ")]
            if len(vals) >= 2:
                nonempty += 1
        out.write(f"  - [{sheet_name}] záznamů: {nonempty}\n")
        if header:
            out.write(f"    header: {[h for h in header if h]}\n")
        file_total += nonempty
    wb.close()
    out.write(f"  ==> celkem v souboru: {file_total}\n")
    totals[path.name] = file_total

out.write("\n### CELKOVÝ SOUHRN\n")
for n, c in totals.items():
    out.write(f"  {n}: {c}\n")
out.write(f"  TOTAL: {sum(totals.values())}\n")

Path(HERE / "_counts.txt").write_text(out.getvalue(), encoding="utf-8")
print("done")
