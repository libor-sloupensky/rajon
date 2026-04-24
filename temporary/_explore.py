"""Průzkum struktury všech xlsx souborů v adresáři."""
import openpyxl
import io
import sys
from pathlib import Path

HERE = Path(__file__).parent
files = sorted(HERE.glob("*.xlsx"))

out = io.StringIO()

for path in files:
    out.write("=" * 80 + "\n")
    out.write(f"FILE: {path.name}  ({path.stat().st_size} B)\n")
    out.write("=" * 80 + "\n")
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        out.write(f"  !! error: {e}\n")
        continue

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out.write(f"\n--- sheet: '{sheet_name}'  rows={ws.max_row} cols={ws.max_column} ---\n")
        preview_rows = 8
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= preview_rows:
                break
            trimmed = [("" if v is None else str(v))[:45] for v in row]
            out.write(f"  r{i+1}: {trimmed}\n")
    wb.close()

Path(HERE / "_structure.txt").write_text(out.getvalue(), encoding="utf-8")
print(f"Written {len(out.getvalue())} chars to _structure.txt")
